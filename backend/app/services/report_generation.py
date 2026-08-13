"""Report generation: shared data-fetching + format-building logic
used by BOTH manual downloads (reports.py's Excel/JSON/CSV/PDF
endpoints) and the scheduled-email job (see services/scheduler.py's
scheduled-reports job). One implementation, not two -- a manual CSV
download and a scheduled CSV email must never be able to drift apart
in what columns they include.
"""
import io
import csv
import json

from ..models.prediction import Prediction, TripSimulation
from ..models.ev_vehicle import EVVehicle

PREDICTION_COLUMNS = [
    ('Date', lambda p: p.created_at.strftime('%Y-%m-%d %H:%M') if p.created_at else ''),
    ('Vehicle', lambda p: _vehicle_name(p.vehicle_id)),
    ('Temperature (C)', lambda p: p.temperature_c),
    ('Humidity (%)', lambda p: p.humidity),
    ('Wind Speed (km/h)', lambda p: p.wind_speed_kmh),
    ('Battery %', lambda p: p.battery_percentage),
    ('Speed (km/h)', lambda p: p.vehicle_speed_kmh),
    ('HVAC', lambda p: 'Yes' if p.hvac_usage else 'No'),
    ('Range Degradation (%)', lambda p: p.range_degradation_pct),
    ('Predicted Range (km)', lambda p: p.predicted_range_km),
    ('Energy (Wh/km)', lambda p: p.energy_consumption_wh_km),
    ('Charging Slowdown (%)', lambda p: p.charging_slowdown_pct),
    ('ML Model', lambda p: p.ml_model_used),
]

TRIP_COLUMNS = [
    ('Date', lambda t: t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else ''),
    ('Source', lambda t: t.source_location),
    ('Destination', lambda t: t.destination),
    ('Distance (km)', lambda t: t.distance_km),
    ('Temperature (C)', lambda t: t.temperature_c),
    ('Speed (km/h)', lambda t: t.speed_kmh),
    ('Heater', lambda t: 'Yes' if t.heater_usage else 'No'),
    ('Battery Usage (%)', lambda t: t.estimated_battery_usage_pct),
    ('Remaining Range (km)', lambda t: t.predicted_remaining_range_km),
    ('Charging Stops', lambda t: t.charging_stops_required),
    ('Arrival Battery (%)', lambda t: t.estimated_arrival_battery_pct),
]


def _vehicle_name(vehicle_id):
    vehicle = EVVehicle.query.get(vehicle_id)
    return f"{vehicle.manufacturer} {vehicle.model_name}" if vehicle else 'Unknown'


def get_report_rows(user_id, report_type, limit=None):
    """Returns (headers, rows): headers is a list of column-name
    strings, rows is a list of lists of already-formatted values, in
    column order -- the one shape every export format below consumes.
    'summary' aggregates a user's activity (via services/analytics.py)
    into metric/value pairs rather than a raw row dump.
    """
    if report_type == 'predictions':
        query = Prediction.query.filter_by(user_id=user_id).order_by(Prediction.created_at.desc())
        if limit:
            query = query.limit(limit)
        items = query.all()
        headers = [c[0] for c in PREDICTION_COLUMNS]
        rows = [[fn(p) for _, fn in PREDICTION_COLUMNS] for p in items]
        return headers, rows

    if report_type == 'trips':
        query = TripSimulation.query.filter_by(user_id=user_id).order_by(TripSimulation.created_at.desc())
        if limit:
            query = query.limit(limit)
        items = query.all()
        headers = [c[0] for c in TRIP_COLUMNS]
        rows = [[fn(t) for _, fn in TRIP_COLUMNS] for t in items]
        return headers, rows

    if report_type == 'summary':
        from . import analytics
        stats = analytics.user_analytics(user_id) or {}
        headers = ['Metric', 'Value']
        rows = [[k, v] for k, v in stats.items()]
        return headers, rows

    raise ValueError(f"Unknown report_type '{report_type}'")


def build_csv(headers, rows):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    return output.getvalue().encode('utf-8')


def build_json(headers, rows):
    data = [dict(zip(headers, row)) for row in rows]
    return json.dumps(data, indent=2, default=str).encode('utf-8')


def build_xlsx(headers, rows, sheet_title='Report'):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel's own sheet-name length limit

    header_fill = PatternFill(start_color='1A73E8', end_color='1A73E8', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = max(12, len(str(header)) + 2)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_pdf(headers, rows, title, subtitle=None, max_rows=50):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles['Title']), Spacer(1, 12)]
    if subtitle:
        elements += [Paragraph(subtitle, styles['Normal']), Spacer(1, 20)]

    if rows:
        table_data = [headers] + [[str(v) if v is not None else '' for v in row] for row in rows[:max_rows]]
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a73e8')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4ff')]),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph('No data available for this report.', styles['Normal']))

    doc.build(elements)
    return buffer.getvalue()


FORMAT_MIME_TYPES = {
    'csv': 'text/csv',
    'json': 'application/json',
    'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'pdf': 'application/pdf',
}


def generate_report_bytes(user_id, report_type, format, title=None, subtitle=None, limit=None):
    """One call: fetch rows + build the requested format's bytes.
    Returns (file_bytes, row_count, mime_type).
    """
    if format not in FORMAT_MIME_TYPES:
        raise ValueError(f"Unsupported format '{format}'")

    headers, rows = get_report_rows(user_id, report_type, limit=limit)

    if format == 'csv':
        file_bytes = build_csv(headers, rows)
    elif format == 'json':
        file_bytes = build_json(headers, rows)
    elif format == 'xlsx':
        file_bytes = build_xlsx(headers, rows, sheet_title=report_type.title())
    else:  # pdf
        file_bytes = build_pdf(headers, rows, title or f"{report_type.title()} Report", subtitle)

    return file_bytes, len(rows), FORMAT_MIME_TYPES[format]
