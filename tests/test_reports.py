"""Flask + DB integration tests for the Reports phase: Excel/JSON
export, Scheduled Reports CRUD + send-now, Report History, and the
Printable Dashboard. Same skip-if-flask_sqlalchemy-missing convention
as the other DB-backed test files.
"""
import pytest
from datetime import datetime, timedelta

flask_sqlalchemy = pytest.importorskip("flask_sqlalchemy", reason="flask_sqlalchemy not installed in this environment")

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'backend'))

from app import create_app, db
from app.models.user import User
from app.models.ev_vehicle import EVVehicle
from app.models.prediction import Prediction, TripSimulation
from app.models.report import ReportSchedule, ReportHistory
from app.services import report_generation


@pytest.fixture
def app():
    app = create_app('testing')
    with app.app_context():
        db.create_all()
        yield app
        db.session.remove()
        db.drop_all()


@pytest.fixture
def user_and_vehicle(app):
    with app.app_context():
        user = User(username='reportuser', email='reportuser@example.com')
        user.set_password('pass12345')
        vehicle = EVVehicle(
            model_name='Test Model', manufacturer='TestCo', battery_capacity_kwh=75,
            epa_range_km=400, vehicle_weight_kg=1900, battery_chemistry='NMC',
            charging_type='CCS', max_charging_power_kw=150, drivetrain='AWD', year=2024,
        )
        db.session.add_all([user, vehicle])
        db.session.commit()
        return user.id, vehicle.id


@pytest.fixture
def client(app):
    return app.test_client()


@pytest.fixture
def auth_client(client, app, user_and_vehicle):
    client.post('/login', data={'username': 'reportuser', 'password': 'pass12345'})
    return client


def _add_prediction(app, user_id, vehicle_id):
    with app.app_context():
        db.session.add(Prediction(
            user_id=user_id, vehicle_id=vehicle_id, temperature_c=-10, humidity=50,
            wind_speed_kmh=10, precipitation='none', battery_percentage=80,
            vehicle_speed_kmh=70, hvac_usage=True, terrain_type='flat', battery_age_years=1,
            range_degradation_pct=20.0, predicted_range_km=300, energy_consumption_wh_km=190,
            charging_slowdown_pct=15.0, ml_model_used='random_forest', prediction_confidence=0.85,
        ))
        db.session.commit()


def _add_trip(app, user_id, vehicle_id):
    with app.app_context():
        db.session.add(TripSimulation(
            user_id=user_id, vehicle_id=vehicle_id, source_location='A', destination='B',
            distance_km=100, temperature_c=-5, speed_kmh=80, heater_usage=True,
            estimated_battery_usage_pct=25, predicted_remaining_range_km=250,
            charging_stops_required=0, estimated_arrival_battery_pct=55,
        ))
        db.session.commit()


# --- services/report_generation.py direct tests ---

def test_get_report_rows_predictions(app, user_and_vehicle):
    user_id, vehicle_id = user_and_vehicle
    _add_prediction(app, user_id, vehicle_id)
    with app.app_context():
        headers, rows = report_generation.get_report_rows(user_id, 'predictions')
    assert 'Range Degradation (%)' in headers
    assert len(rows) == 1
    assert rows[0][headers.index('Vehicle')] == 'TestCo Test Model'


def test_get_report_rows_trips(app, user_and_vehicle):
    user_id, vehicle_id = user_and_vehicle
    _add_trip(app, user_id, vehicle_id)
    with app.app_context():
        headers, rows = report_generation.get_report_rows(user_id, 'trips')
    assert 'Distance (km)' in headers
    assert len(rows) == 1


def test_get_report_rows_summary(app, user_and_vehicle):
    user_id, vehicle_id = user_and_vehicle
    _add_prediction(app, user_id, vehicle_id)
    with app.app_context():
        headers, rows = report_generation.get_report_rows(user_id, 'summary')
    assert headers == ['Metric', 'Value']
    metrics = {row[0]: row[1] for row in rows}
    assert metrics['total_predictions'] == 1


def test_get_report_rows_unknown_type_raises(app, user_and_vehicle):
    user_id, _ = user_and_vehicle
    with app.app_context():
        with pytest.raises(ValueError):
            report_generation.get_report_rows(user_id, 'bogus')


def test_build_csv_produces_valid_content():
    csv_bytes = report_generation.build_csv(['A', 'B'], [[1, 2], [3, 4]])
    text = csv_bytes.decode('utf-8')
    assert 'A,B' in text
    assert '1,2' in text


def test_build_json_produces_valid_content():
    import json
    json_bytes = report_generation.build_json(['A', 'B'], [[1, 2]])
    data = json.loads(json_bytes)
    assert data == [{'A': 1, 'B': 2}]


def test_build_xlsx_produces_valid_workbook():
    from openpyxl import load_workbook
    xlsx_bytes = report_generation.build_xlsx(['A', 'B'], [[1, 2], [3, 4]], sheet_title='Test')
    wb = load_workbook(io_bytes(xlsx_bytes))
    ws = wb.active
    assert ws.title == 'Test'
    assert ws.cell(row=1, column=1).value == 'A'
    assert ws.cell(row=2, column=1).value == 1


def io_bytes(b):
    import io
    return io.BytesIO(b)


def test_generate_report_bytes_end_to_end(app, user_and_vehicle):
    user_id, vehicle_id = user_and_vehicle
    _add_prediction(app, user_id, vehicle_id)
    with app.app_context():
        file_bytes, row_count, mime = report_generation.generate_report_bytes(user_id, 'predictions', 'csv')
    assert row_count == 1
    assert mime == 'text/csv'
    assert len(file_bytes) > 0


def test_generate_report_bytes_unsupported_format_raises(app, user_and_vehicle):
    user_id, _ = user_and_vehicle
    with app.app_context():
        with pytest.raises(ValueError):
            report_generation.generate_report_bytes(user_id, 'predictions', 'bogus')


# --- ReportSchedule.is_due() ---

def test_schedule_is_due_when_never_sent(app, user_and_vehicle):
    user_id, _ = user_and_vehicle
    with app.app_context():
        s = ReportSchedule(user_id=user_id, name='Test', report_type='predictions', format='csv', frequency='daily')
        assert s.is_due() is True


def test_schedule_not_due_when_disabled(app, user_and_vehicle):
    user_id, _ = user_and_vehicle
    with app.app_context():
        s = ReportSchedule(user_id=user_id, name='Test', report_type='predictions', format='csv', frequency='daily', enabled=False)
        assert s.is_due() is False


def test_schedule_not_due_within_frequency_window(app, user_and_vehicle):
    user_id, _ = user_and_vehicle
    with app.app_context():
        s = ReportSchedule(user_id=user_id, name='Test', report_type='predictions', format='csv', frequency='weekly')
        s.last_sent_at = datetime.utcnow() - timedelta(days=1)
        assert s.is_due() is False


def test_schedule_due_after_frequency_window(app, user_and_vehicle):
    user_id, _ = user_and_vehicle
    with app.app_context():
        s = ReportSchedule(user_id=user_id, name='Test', report_type='predictions', format='csv', frequency='daily')
        s.last_sent_at = datetime.utcnow() - timedelta(days=2)
        assert s.is_due() is True


# --- services/scheduled_reports.py ---

def test_run_schedule_logs_history_and_updates_last_sent(app, user_and_vehicle):
    from app.services.scheduled_reports import run_schedule
    user_id, vehicle_id = user_and_vehicle
    _add_prediction(app, user_id, vehicle_id)
    with app.app_context():
        s = ReportSchedule(user_id=user_id, name='Test', report_type='predictions', format='csv', frequency='daily')
        db.session.add(s)
        db.session.commit()
        history = run_schedule(app.config, s)
        assert history is not None
        assert history.source == 'scheduled'
        assert history.email_status == 'mail not configured'  # test config has no MAIL_USERNAME
        assert history.delivered_via_email is False
        assert s.last_sent_at is not None


def test_run_due_report_schedules_only_sends_due_ones(app, user_and_vehicle):
    from app.services.scheduled_reports import run_due_report_schedules
    user_id, vehicle_id = user_and_vehicle
    with app.app_context():
        due = ReportSchedule(user_id=user_id, name='Due', report_type='predictions', format='csv', frequency='daily')
        not_due = ReportSchedule(user_id=user_id, name='NotDue', report_type='predictions', format='csv', frequency='weekly')
        not_due.last_sent_at = datetime.utcnow()
        db.session.add_all([due, not_due])
        db.session.commit()
        results = run_due_report_schedules(app)
    assert results['due'] == 1
    assert results['checked'] == 2


# --- API endpoint tests ---

def test_reports_index_requires_login(client):
    resp = client.get('/reports/')
    assert resp.status_code in (302, 401)


def test_reports_index_loads(auth_client):
    resp = auth_client.get('/reports/')
    assert resp.status_code == 200


def test_printable_dashboard_loads(auth_client):
    resp = auth_client.get('/reports/printable')
    assert resp.status_code == 200


def test_export_excel_predictions(auth_client, user_and_vehicle, app):
    user_id, vehicle_id = user_and_vehicle
    _add_prediction(app, user_id, vehicle_id)
    resp = auth_client.get('/reports/api/excel/predictions')
    assert resp.status_code == 200
    assert resp.headers['Content-Type'].startswith('application/vnd.openxmlformats')


def test_export_excel_invalid_type_400s(auth_client):
    resp = auth_client.get('/reports/api/excel/bogus')
    assert resp.status_code == 400


def test_export_json_predictions(auth_client, user_and_vehicle, app):
    user_id, vehicle_id = user_and_vehicle
    _add_prediction(app, user_id, vehicle_id)
    resp = auth_client.get('/reports/api/json/predictions')
    assert resp.status_code == 200
    data = resp.get_json()
    assert len(data) == 1


def test_export_json_summary(auth_client, user_and_vehicle, app):
    user_id, vehicle_id = user_and_vehicle
    _add_prediction(app, user_id, vehicle_id)
    resp = auth_client.get('/reports/api/json/summary')
    assert resp.status_code == 200


def test_history_records_manual_exports(auth_client, user_and_vehicle, app):
    user_id, vehicle_id = user_and_vehicle
    _add_prediction(app, user_id, vehicle_id)
    auth_client.get('/reports/api/excel/predictions')
    auth_client.get('/reports/api/json/predictions')
    resp = auth_client.get('/reports/api/history')
    assert resp.status_code == 200
    history = resp.get_json()
    assert len(history) == 2
    assert {h['format'] for h in history} == {'xlsx', 'json'}
    assert all(h['source'] == 'manual' for h in history)


def test_schedule_crud_flow(auth_client):
    resp = auth_client.post('/reports/api/schedules', json={
        'name': 'Weekly predictions', 'report_type': 'predictions', 'format': 'csv', 'frequency': 'weekly',
    })
    assert resp.status_code == 201
    schedule = resp.get_json()
    assert schedule['enabled'] is True
    schedule_id = schedule['id']

    resp = auth_client.get('/reports/api/schedules')
    assert len(resp.get_json()) == 1

    resp = auth_client.patch(f'/reports/api/schedules/{schedule_id}', json={'enabled': False})
    assert resp.status_code == 200
    assert resp.get_json()['enabled'] is False

    resp = auth_client.delete(f'/reports/api/schedules/{schedule_id}')
    assert resp.status_code == 200
    assert auth_client.get('/reports/api/schedules').get_json() == []


def test_schedule_creation_validates_report_type(auth_client):
    resp = auth_client.post('/reports/api/schedules', json={'name': 'X', 'report_type': 'bogus'})
    assert resp.status_code == 400


def test_schedule_creation_validates_format(auth_client):
    resp = auth_client.post('/reports/api/schedules', json={'name': 'X', 'report_type': 'predictions', 'format': 'bogus'})
    assert resp.status_code == 400


def test_schedule_creation_validates_frequency(auth_client):
    resp = auth_client.post('/reports/api/schedules', json={'name': 'X', 'report_type': 'predictions', 'format': 'csv', 'frequency': 'bogus'})
    assert resp.status_code == 400


def test_schedule_creation_requires_name(auth_client):
    resp = auth_client.post('/reports/api/schedules', json={'report_type': 'predictions'})
    assert resp.status_code == 400


def test_send_schedule_now_logs_history(auth_client, user_and_vehicle, app):
    user_id, vehicle_id = user_and_vehicle
    _add_prediction(app, user_id, vehicle_id)
    resp = auth_client.post('/reports/api/schedules', json={
        'name': 'Test schedule', 'report_type': 'predictions', 'format': 'csv', 'frequency': 'weekly',
    })
    schedule_id = resp.get_json()['id']

    resp = auth_client.post(f'/reports/api/schedules/{schedule_id}/send-now')
    assert resp.status_code == 200
    history_entry = resp.get_json()
    assert history_entry['source'] == 'scheduled'

    history = auth_client.get('/reports/api/history').get_json()
    assert any(h['source'] == 'scheduled' for h in history)


def test_send_schedule_now_unknown_schedule_404s(auth_client):
    resp = auth_client.post('/reports/api/schedules/99999/send-now')
    assert resp.status_code == 404


def test_cannot_delete_someone_elses_schedule(auth_client, app, user_and_vehicle):
    user_id, vehicle_id = user_and_vehicle
    with app.app_context():
        other = User(username='otherreportuser', email='other@example.com')
        other.set_password('pass12345')
        db.session.add(other)
        db.session.commit()
        other_id = other.id
        s = ReportSchedule(user_id=other_id, name='Not yours', report_type='predictions', format='csv', frequency='weekly')
        db.session.add(s)
        db.session.commit()
        sid = s.id

    resp = auth_client.delete(f'/reports/api/schedules/{sid}')
    assert resp.status_code == 404
